import logging
import tempfile
import pandas as pd
import tabula
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .image_processor import is_image_based_pdf, process_image_based_pdf
from typing import Dict, List
import PyPDF2
import re

def clean_amount(amount_str):
    """Clean and format amount strings"""
    if pd.isna(amount_str):
        return ''
    # Remove currency symbols and cleanup
    amount_str = str(amount_str).replace('$', '').replace(',', '').strip()
    # Handle brackets for negative numbers
    if '(' in amount_str and ')' in amount_str:
        amount_str = '-' + amount_str.replace('(', '').replace(')', '')
    try:
        # Try to convert to float to validate
        float(amount_str)
        return amount_str
    except ValueError:
        return ''

def parse_date(date_str):
    """Parse date string from bank statement format"""
    try:
        if not date_str or pd.isna(date_str):
            return None

        date_str = str(date_str).strip().upper()

        # Skip rows that aren't dates
        if any(word in date_str for word in ['TOTALS', 'BALANCE', 'OPENING']):
            return None

        # Handle day and month format (e.g., "26 APR")
        parts = date_str.split()
        if len(parts) == 2:
            try:
                day = int(parts[0])
                month = parts[1][:3]  # Take first 3 chars of month
                current_year = datetime.now().year
                # Handle special case for dates like "31 APR"
                if month == 'APR' and day == 31:
                    day = 30
                date_str = f"{day:02d} {month} {current_year}"
                parsed_date = datetime.strptime(date_str, '%d %b %Y')
                return parsed_date
            except (ValueError, IndexError) as e:
                logging.debug(f"Date parse error: {e} for {date_str}")
                return None

        return None
    except Exception as e:
        logging.debug(f"Failed to parse date: {date_str}, error: {str(e)}")
        return None

def process_transaction_rows(table, page_idx):
    """Process rows and handle multi-line transactions"""
    processed_data = []
    current_buffer = []

    # Clean the table
    table = table.dropna(how='all').reset_index(drop=True)

    logging.debug(f"Starting to process table on page {page_idx} with {len(table)} rows")
    logging.debug(f"Table columns: {table.columns}")
    logging.debug(f"First few rows: {table.head()}")

    def process_buffer():
        if not current_buffer:
            return None

        logging.debug(f"Processing buffer with {len(current_buffer)} rows: {current_buffer}")

        # Get date from first row
        date = parse_date(current_buffer[0][0])
        if not date:
            logging.debug(f"Failed to parse date from: {current_buffer[0][0]}")
            return None

        # Initialize transaction
        transaction = {
            'Date': date.strftime('%d %b'),
            'Transaction Details': '',
            'Withdrawals ($)': '',
            'Deposits ($)': '',
            'Balance ($)': '',
            '_page_idx': page_idx,
            '_row_idx': int(current_buffer[0][-1]),
        }

        # Process all rows
        details = []
        for row in current_buffer:
            # Add description
            if row[1].strip():
                details.append(row[1].strip())
                logging.debug(f"Added description: {row[1].strip()}")

            # Process amounts with detailed logging
            withdrawal = clean_amount(row[2])
            deposit = clean_amount(row[3])
            balance = clean_amount(row[4]) if len(row) > 4 else ''

            logging.debug(f"Processing amounts - W: {withdrawal}, D: {deposit}, B: {balance}")

            # Update amounts if not already set
            if withdrawal and not transaction['Withdrawals ($)']:
                transaction['Withdrawals ($)'] = withdrawal
                logging.debug(f"Set withdrawal: {withdrawal}")
            if deposit and not transaction['Deposits ($)']:
                transaction['Deposits ($)'] = deposit
                logging.debug(f"Set deposit: {deposit}")
            if balance and not transaction['Balance ($)']:
                transaction['Balance ($)'] = balance
                logging.debug(f"Set balance: {balance}")

        # Join details
        transaction['Transaction Details'] = '\n'.join(filter(None, details))
        logging.debug(f"Final transaction: {transaction}")
        return transaction

    # Process each row
    for idx, row in table.iterrows():
        # Clean row values and add index
        row_values = [str(val).strip() if not pd.isna(val) else '' for val in row]
        row_values.append(idx)

        logging.debug(f"Processing row {idx}: {row_values}")

        # Check for date and content
        has_date = bool(parse_date(row_values[0]))
        has_content = any(val.strip() for val in row_values[1:5])

        logging.debug(f"Row analysis - has_date: {has_date}, has_content: {has_content}")

        if has_date:
            # Process previous buffer if exists
            if current_buffer:
                trans = process_buffer()
                if trans:
                    processed_data.append(trans)
                current_buffer = []

            # Start new buffer
            current_buffer = [row_values]
            logging.debug(f"Started new transaction: {row_values}")

        elif current_buffer and has_content:
            # Add to current buffer
            current_buffer.append(row_values)
            logging.debug(f"Added to current transaction: {row_values}")

    # Process final buffer
    if current_buffer:
        trans = process_buffer()
        if trans:
            processed_data.append(trans)

    # Sort by row index
    processed_data.sort(key=lambda x: (x['_page_idx'], x['_row_idx']))

    # Remove tracking fields
    for trans in processed_data:
        trans.pop('_page_idx', None)
        trans.pop('_row_idx', None)

    # Log results
    logging.debug(f"Processed {len(processed_data)} transactions")
    for idx, trans in enumerate(processed_data):
        logging.debug(f"Transaction {idx}: {trans}")

    return processed_data

def is_valid_transaction(transaction: Dict) -> bool:
    """Validate transaction data"""
    try:
        # Allow opening balance entries
        if transaction.get('is_opening_balance') or (
            'OPENING BALANCE' in transaction['Transaction Details'].upper() and
            transaction['Balance ($)']
        ):
            return True

        # Must have date and some content
        if not transaction['Date']:
            return False

        # Must have some details or amounts
        has_content = any([
            transaction['Transaction Details'],
            transaction['Withdrawals ($)'],
            transaction['Deposits ($)'],
            transaction['Balance ($)']
        ])
        if not has_content:
            return False

        # Skip other header/footer rows
        skip_words = ['closing', 'balance brought', 'balance carried', 'total']
        details_lower = transaction['Transaction Details'].lower()
        if any(word in details_lower for word in skip_words):
            return False

        return True
    except Exception:
        return False

def detect_bank_statement_type(pdf_path: str) -> str:
    """Detect the type of bank statement based on content analysis"""
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            first_page_text = pdf_reader.pages[0].extract_text().upper()

            if 'NATIONWIDE' in first_page_text:
                return 'nationwide'
            return 'generic'
    except Exception as e:
        logging.error(f"Error detecting bank statement type: {str(e)}")
        return 'generic'

def process_nationwide_statement(table):
    """Process Nationwide bank statement specific format"""
    try:
        processed_data = []
        logging.debug(f"Processing Nationwide statement table with shape: {table.shape}")
        logging.debug(f"Table columns: {table.columns.tolist()}")
        logging.debug(f"First few rows:\n{table.head()}")

        # Clean and standardize the table
        table = table.dropna(how='all').reset_index(drop=True)

        # Find the header row
        header_row_idx = None
        for idx, row in table.iterrows():
            row_values = [str(val).strip().upper() for val in row if not pd.isna(val)]
            row_text = ' '.join(row_values)
            logging.debug(f"Checking row {idx}: {row_text}")

            if any(keyword in row_text for keyword in ['DATE', 'DESCRIPTION', 'PAYMENTS', 'RECEIPTS', 'BALANCE']):
                header_row_idx = idx
                logging.debug(f"Found header row at index {idx}")
                break

        if header_row_idx is None:
            logging.error("Could not find header row in table")
            return []

        # Set the header and clean the table
        table.columns = table.iloc[header_row_idx]
        table = table.iloc[header_row_idx + 1:].reset_index(drop=True)

        # Map columns to standardized names
        column_mapping = {}
        for col in table.columns:
            col_str = str(col).upper()
            if 'DATE' in col_str:
                column_mapping[col] = 'Date'
            elif any(x in col_str for x in ['DESCRIPTION', 'DETAILS', 'TRANSACTION']):
                column_mapping[col] = 'Description'
            elif any(x in col_str for x in ['PAYMENT', 'OUT', 'DEBIT', 'WITHDRAWALS']):
                column_mapping[col] = 'Withdrawals'
            elif any(x in col_str for x in ['RECEIPT', 'IN', 'CREDIT', 'DEPOSITS']):
                column_mapping[col] = 'Deposits'
            elif 'BALANCE' in col_str:
                column_mapping[col] = 'Balance'

        logging.debug(f"Column mapping: {column_mapping}")

        # Rename columns using the mapping
        table = table.rename(columns=column_mapping)
        required_columns = ['Date', 'Description', 'Withdrawals', 'Deposits', 'Balance']

        # Verify we have all required columns
        missing_columns = [col for col in required_columns if col not in table.columns]
        if missing_columns:
            logging.error(f"Missing required columns: {missing_columns}")
            return []

        # Process each row
        for idx, row in table.iterrows():
            try:
                # Skip rows without any transaction data
                if row.isna().all():
                    continue

                # Clean and format the data
                date = str(row['Date']).strip()
                details = str(row['Description']).strip()
                withdrawal = clean_amount(str(row['Withdrawals']))
                deposit = clean_amount(str(row['Deposits']))
                balance = clean_amount(str(row['Balance']))

                # Skip non-transaction rows
                if not date or not details:
                    continue

                # Create transaction record
                transaction = {
                    'Date': date,
                    'Transaction Details': details,
                    'Withdrawals ($)': withdrawal,
                    'Deposits ($)': deposit,
                    'Balance ($)': balance
                }

                if is_valid_transaction(transaction):
                    processed_data.append(transaction)
                    logging.debug(f"Added transaction: {transaction}")

            except Exception as e:
                logging.error(f"Error processing row {idx}: {str(e)}")
                continue

        logging.info(f"Successfully processed {len(processed_data)} transactions")
        return processed_data

    except Exception as e:
        logging.error(f"Error processing Nationwide statement: {str(e)}")
        return []

def extract_text_from_area(pdf_path: str, selected_area: Dict) -> str:
    """Extract text from a specific area of a PDF page"""
    try:
        logging.info(f"Extracting text from selected area: {selected_area}")

        # Use tabula to extract tables from the specific area
        area = [
            selected_area['y'] * 100,  # top
            selected_area['x'] * 100,  # left
            (selected_area['y'] + selected_area['height']) * 100,  # bottom
            (selected_area['x'] + selected_area['width']) * 100  # right
        ]

        tables = tabula.read_pdf(
            pdf_path,
            pages=selected_area.get('page', 1),
            area=area,
            relative_area=True,
            stream=True,
            guess=False,
            pandas_options={'header': None}
        )

        if not tables:
            return ""

        # Convert table to text
        text = ""
        for table in tables:
            text += table.to_string(index=False, header=False) + "\n"

        logging.debug(f"Extracted text from area:\n{text}")
        return text
    except Exception as e:
        logging.error(f"Error extracting text from area: {str(e)}")
        return ""

def parse_text_to_transactions(text: str) -> List[Dict]:
    """Parse extracted text into transactions"""
    try:
        transactions = []
        current_transaction = None
        lines = text.split('\n')

        # Date pattern matching
        date_pattern = re.compile(r'(\d{1,2})\s*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)', re.IGNORECASE)
        amount_pattern = re.compile(r'\$?\s*-?\d+(?:,\d{3})*(?:\.\d{2})?')

        for line in lines:
            line = line.strip()
            if not line:
                continue

            logging.debug(f"Processing line: {line}")

            # Check for date at start of line
            date_match = date_pattern.search(line)

            if date_match:
                # If we found a date, start a new transaction
                if current_transaction and is_valid_transaction(current_transaction):
                    transactions.append(current_transaction)

                # Initialize new transaction
                current_transaction = {
                    'Date': date_match.group().strip(),
                    'Transaction Details': line[date_match.end():].strip(),
                    'Withdrawals ($)': '',
                    'Deposits ($)': '',
                    'Balance ($)': ''
                }

                # Look for amounts in the rest of the line
                amounts = amount_pattern.findall(line[date_match.end():])
                if amounts:
                    for amount in amounts:
                        amount = clean_amount(amount)
                        if amount:
                            if float(amount) < 0:
                                current_transaction['Withdrawals ($)'] = str(abs(float(amount)))
                            else:
                                current_transaction['Deposits ($)'] = amount

            elif current_transaction:
                # Continue with current transaction
                amounts = amount_pattern.findall(line)
                details = line

                # Remove amount strings from details
                for amount in amounts:
                    details = details.replace(amount, '').strip()

                if details:
                    current_transaction['Transaction Details'] += ' ' + details

                # Process amounts
                if amounts and not any([current_transaction['Withdrawals ($)'], 
                                      current_transaction['Deposits ($)'], 
                                      current_transaction['Balance ($)']]):
                    for amount in amounts:
                        amount = clean_amount(amount)
                        if amount:
                            if float(amount) < 0:
                                current_transaction['Withdrawals ($)'] = str(abs(float(amount)))
                            else:
                                current_transaction['Deposits ($)'] = amount

        # Add the last transaction
        if current_transaction and is_valid_transaction(current_transaction):
            transactions.append(current_transaction)

        logging.info(f"Parsed {len(transactions)} transactions from text")
        return transactions

    except Exception as e:
        logging.error(f"Error parsing text to transactions: {str(e)}")
        return []

def extract_tables_from_pdf(pdf_path, selected_areas=None, java_options=None):
    """Extract tables from PDF using both lattice and stream methods"""
    try:
        logging.info(f"Starting table extraction from {pdf_path}")

        # Get PDF dimensions using PyPDF2
        with open(pdf_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            first_page = pdf_reader.pages[0]
            pdf_width = float(first_page.mediabox.width)
            pdf_height = float(first_page.mediabox.height)
            num_pages = len(pdf_reader.pages)
            logging.info(f"PDF has {num_pages} pages, dimensions: {pdf_width}x{pdf_height}")

        all_tables = []

        # Process each page
        for page_num in range(1, num_pages + 1):
            logging.debug(f"Processing page {page_num}")

            page_areas = None
            if selected_areas:
                # Filter areas for current page
                page_areas = [
                    [
                        area['y'] * 100,  # top
                        area['x'] * 100,  # left
                        (area['y'] + area['height']) * 100,  # bottom
                        (area['x'] + area['width']) * 100  # right
                    ]
                    for area in selected_areas if area.get('page', 1) == page_num
                ]
                if not page_areas:
                    logging.debug(f"No selected areas for page {page_num}")
                    continue
                logging.debug(f"Found areas for page {page_num}: {page_areas}")

            # Try extraction methods
            methods = [
                {'lattice': True, 'stream': False},
                {'lattice': False, 'stream': True},
                {'lattice': True, 'stream': True}
            ]

            page_tables = []
            for method in methods:
                try:
                    logging.debug(f"Trying extraction with method: {method}")
                    tables = tabula.read_pdf(
                        pdf_path,
                        pages=str(page_num),
                        multiple_tables=True,
                        guess=True,
                        area=page_areas[0] if page_areas else None,
                        relative_area=False if page_areas else True,
                        lattice=method['lattice'],
                        stream=method['stream'],
                        pandas_options={'header': None},
                        java_options=java_options
                    )

                    if tables:
                        logging.debug(f"Found {len(tables)} tables with method {method}")
                        page_tables.extend(tables)

                except Exception as e:
                    logging.error(f"Error with method {method}: {str(e)}")
                    continue

            if page_tables:
                # Add page information to tables
                for table in page_tables:
                    table.attrs = {'page_number': page_num}
                    logging.debug(f"Table shape: {table.shape}")
                    logging.debug(f"Table preview:\n{table.head()}")
                all_tables.extend(page_tables)

        if not all_tables:
            logging.error("No tables could be extracted from any page")
            return []

        logging.info(f"Successfully extracted {len(all_tables)} tables total")
        return all_tables

    except Exception as e:
        logging.error(f"Error in table extraction: {str(e)}")
        return []

def convert_pdf_to_data(pdf_path: str, selected_areas=None):
    """Extract data from PDF bank statement and return as list of dictionaries"""
    try:
        logging.info(f"Starting data extraction from {pdf_path}")

        if not os.path.exists(pdf_path):
            logging.error("PDF file not found")
            return None

        all_transactions = []

        if selected_areas:
            # Process each selected area
            for area in selected_areas:
                # Extract text from the selected area
                text = extract_text_from_area(pdf_path, area)
                if text:
                    # Parse the extracted text into transactions
                    transactions = parse_text_to_transactions(text)
                    if transactions:
                        all_transactions.extend(transactions)

        if not all_transactions:
            logging.warning("No transactions found in selected areas, trying full page extraction")
            # Fallback to original table extraction method
            tables = extract_tables_from_pdf(pdf_path, selected_areas)
            if tables:
                for table in tables:
                    if len(table.columns) >= 4:
                        table.columns = range(len(table.columns))
                        transactions = process_transaction_rows(table, 1)
                        all_transactions.extend([t for t in transactions if is_valid_transaction(t)])

        if not all_transactions:
            logging.error("No valid transactions could be extracted")
            return None

        logging.info(f"Successfully extracted {len(all_transactions)} transactions")
        return all_transactions

    except Exception as e:
        logging.error(f"Error in data extraction: {str(e)}")
        return None

def convert_pdf(pdf_path: str, output_format: str = 'excel', selected_areas=None):
    """Convert PDF bank statement to Excel/CSV"""
    try:
        # Extract data using the improved processing logic
        processed_data = convert_pdf_to_data(pdf_path, selected_areas)

        if not processed_data:
            return None

        # Convert to DataFrame
        df = pd.DataFrame(processed_data)

        # Create output file
        temp_file = tempfile.NamedTemporaryFile(delete=False)

        if output_format == 'excel':
            output_path = f"{temp_file.name}.xlsx"
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Transactions')
                workbook = writer.book
                worksheet = writer.sheets['Transactions']

                # Format headers
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                header_alignment = Alignment(horizontal='center')

                for col in range(len(df.columns)):
                    cell = worksheet.cell(row=1, column=col+1)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # Adjust column widths
                for idx, column in enumerate(worksheet.columns, 1):
                    max_length = max(len(str(cell.value)) for cell in column)
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

                # Set wrap text for transaction details
                for cell in worksheet['B']:
                    cell.alignment = Alignment(wrapText=True)
        else:
            output_path = f"{temp_file.name}.csv"
            df.to_csv(output_path, index=False)

        logging.info(f"Successfully created output file: {output_path}")
        return output_path

    except Exception as e:
        logging.error(f"Error in conversion: {str(e)}")
        return None